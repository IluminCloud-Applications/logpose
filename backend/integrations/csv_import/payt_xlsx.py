import io
import logging
from openpyxl import load_workbook
from integrations.csv_import.schemas import ImportRow

logger = logging.getLogger(__name__)

PAYT_STATUS_MAP = {
    "compra aprovada": "approved",
    "compra reembolsada": "refunded",
    "reembolsado": "refunded",
    "chargeback": "chargeback",
}


def _clean(val) -> str:
    """Limpa valores do XLSX da PayT (trata '-' como vazio)."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.replace(" ", "") == "-":
        return ""
    return s


def _parse_brl(val) -> float:
    """Converte formato BRL (177,30 ou R$ 165,68) para float."""
    s = _clean(val)
    if not s:
        return 0.0
    s = s.replace("R$", "").replace(" ", "").strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _read_sheet(content: bytes) -> tuple[list, list[dict]]:
    """Lê um XLSX e retorna header + lista de dicts."""
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], []

    header = rows[0]
    data = [dict(zip(header, row)) for row in rows[1:]]
    return header, data


def parse_payt_xlsx(
    vendas_content: bytes,
    origem_content: bytes,
) -> list[ImportRow]:
    """
    Parseia os 2 XLSX da PayT:
    - vendas: dados da transação + cliente
    - origem: UTMs de rastreamento
    Cruza pelo Código da Venda.
    """
    _, vendas_list = _read_sheet(vendas_content)
    _, origem_list = _read_sheet(origem_content)

    # Indexar origem pelo código da venda
    origem_map: dict[str, dict] = {}
    for row in origem_list:
        code = _clean(row.get("Código da Venda"))
        if code:
            origem_map[code] = row

    rows: list[ImportRow] = []
    for venda in vendas_list:
        code = _clean(venda.get("Código"))
        if not code:
            continue

        product_name = _clean(venda.get("Produto"))
        if not product_name:
            continue

        origem = origem_map.get(code, {})
        status_raw = _clean(venda.get("Status Compra")).lower()
        status = PAYT_STATUS_MAP.get(status_raw, "pending")

        phone = _clean(venda.get("Telefone"))
        if phone and not phone.startswith("+"):
            phone = f"+55{phone}"

        src = (
            _clean(origem.get("Source (URL)"))
            or _clean(venda.get("Source / Venda Manual"))
            or None
        )

        # Data: vendas usa "10/03/2026 13:36:46", origem usa "10/03/2026 - 13:36:46"
        date_str = _clean(venda.get("Data"))
        if not date_str:
            date_str = _clean(origem.get("Data"))
        date_str = date_str.replace(" - ", " ") if date_str else None

        rows.append(ImportRow(
            external_id=code,
            status=status,
            product_name=product_name,
            product_external_id=_clean(venda.get("Sku")) or product_name,
            product_ticket=_parse_brl(venda.get("Preço do Produto")),
            amount=_parse_brl(venda.get("Você Recebe")),
            customer_name=_clean(venda.get("Cliente")) or None,
            customer_email=_clean(venda.get("Email")),
            customer_cpf=_clean(venda.get("Documento")) or None,
            customer_phone=phone or None,
            checkout_name=_clean(venda.get("Nome do Checkout")) or None,
            checkout_code=_clean(venda.get("Código do Checkout")) or None,
            utm_source=_clean(origem.get("Utm Source (URL)")) or None,
            utm_medium=_clean(origem.get("Utm Medium (URL)")) or None,
            utm_campaign=_clean(origem.get("Utm Campaign (URL)")) or None,
            utm_content=_clean(origem.get("Utm Content (URL)")) or None,
            utm_term=_clean(origem.get("Utm Term (URL)")) or None,
            src=src,
            created_at=date_str,
        ))

    logger.info(f"PayT XLSX: {len(rows)} linhas parseadas ({len(origem_map)} origens)")
    return rows
